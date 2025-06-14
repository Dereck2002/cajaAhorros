from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.db.models.functions import TruncMonth
from django.db.models import Count, Sum
from django.contrib import messages
from .models import Movimiento, Socio, Cargo, Prestamo, PagoPrestamo, Configuracion, GastosAdministrativos
from .forms import SocioForm, PrestamoForm, ConfiguracionForm, GastoAdministrativoForm
from datetime import date, datetime
from dateutil.relativedelta import relativedelta
from django.core.paginator import Paginator
from decimal import Decimal
from django.utils import timezone
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from django.http import HttpResponse
from openpyxl import Workbook
from django.utils.timezone import now
from openpyxl.styles import Font

# Importamos el decorador de roles que creamos
from .decorators import role_required

# --- Función Auxiliar para obtener el Rol ---
def get_user_role(user):
    if user.is_superuser:
        return 'Administrador' # El superusuario siempre es Administrador
    # Devuelve el nombre del primer grupo del usuario, o None si no pertenece a ninguno.
    return user.groups.first().name if user.groups.exists() else None
@login_required
def dashboard(request):
    """
    Esta vista actúa como un centro de redirección después del login.
    Dirige a cada usuario a su página principal correspondiente.
    """
    role = get_user_role(request.user)

    if role == 'Secretaria':
        return redirect('socio_list') # La secretaria va a la lista de socios.
    
    if role == 'Tesorero':
        return redirect('prestamo_list') # El tesorero va a la lista de préstamos.
    
    # El Presidente, Administrador y Superusuario van a la lista de socios como página principal.
    # No se muestra un panel de control, sino que se les dirige a la vista principal de la app.
    return redirect('socio_list')


# --- Vistas de Socios ---
@login_required
@role_required(allowed_roles=['Presidente', 'Secretaria'])
def socio_list(request):
    filtro = request.GET.get('filtro', 'todos')
    socios = Socio.objects.filter(activo=True)
    cargos = Cargo.objects.all()
    hoy = date.today().replace(day=1)

    def meses_faltantes(socio):
        movimientos = Movimiento.objects.filter(socio=socio, entrada__gt=0).order_by('fecha_movimiento')
        if not movimientos.exists():
            return True
        fechas_aportes = movimientos.dates('fecha_movimiento', 'month')
        meses_aporte_set = set((d.year, d.month) for d in fechas_aportes)
        
        # Corrección: Asegurarse de que `movimientos.first()` no sea None
        if not movimientos:
            return True
            
        inicio = movimientos.first().fecha_movimiento.replace(day=1)
        actual = inicio
        while actual <= hoy:
            if (actual.year, actual.month) not in meses_aporte_set:
                return True
            actual += relativedelta(months=1)
        return False

    socios_filtrados = []
    for socio in socios:
        faltante = meses_faltantes(socio)
        if filtro == 'todos' or (filtro == 'al_dia' and not faltante) or (filtro == 'deudores' and faltante):
            socios_filtrados.append(socio)

    paginator = Paginator(socios_filtrados, 10)
    page = request.GET.get('page')
    socios_paginados = paginator.get_page(page)

    role = get_user_role(request.user)
    is_read_only = role == 'Presidente'
    
    context = {
        'socios': socios_paginados,
        'filtro': filtro,
        'cargos': cargos,
        'is_read_only': is_read_only,
        'role': role,
    }
    return render(request, 'socios/socio_list.html', context)

@login_required
@role_required(allowed_roles=['Secretaria'])
def crear_socio(request):
    if request.method == 'POST':
        form = SocioForm(request.POST, request.FILES)
        if form.is_valid():
            socio = form.save()

            # Obtener configuración general
            config = Configuracion.objects.first()
            aporte_inicial = config.aporte_inicial if config else Decimal('0.00')
            gasto_adm = config.gastos_adm if config else Decimal('0.00')
            fecha_hoy = now().date()

            # Registrar el primer aporte como movimiento
            ultimo_mov = Movimiento.objects.filter(socio=socio).order_by('-fecha_movimiento').first()
            saldo_anterior = ultimo_mov.saldo if ultimo_mov else Decimal('0.00')
            nuevo_saldo = saldo_anterior + aporte_inicial

            Movimiento.objects.create(
                socio=socio,
                detalle_movimiento="Aporte inicial",
                entrada=aporte_inicial,
                salida=Decimal('0.00'),
                saldo=nuevo_saldo,
                fecha_movimiento=fecha_hoy
            )

            # Registrar gasto administrativo
            saldo_gasto = GastosAdministrativos.objects.order_by('-fecha').first()
            saldo_anterior_gasto = saldo_gasto.saldo if saldo_gasto else Decimal('0.00')
            nuevo_saldo_gasto = saldo_anterior_gasto + gasto_adm

            GastosAdministrativos.objects.create(
                fecha = socio.fecha_ingreso,
                descripcion="Ingreso por nuevo socio: " + socio.nombre + " " + socio.apellido,
                entrada=gasto_adm,
                salida=Decimal('0.00'),
                saldo=nuevo_saldo_gasto
            )

            return redirect('socio_list')
    else:
        form = SocioForm()
    return render(request, 'socios/crear_socio.html', {'form': form})

@login_required
@role_required(allowed_roles=['Secretaria'])
def eliminar_socio(request, pk):
    socio = get_object_or_404(Socio, pk=pk)

    # Verificar si el socio tiene préstamos como titular o garante que NO estén terminados
    prestamos_como_titular = Prestamo.objects.filter(socio=socio).exclude(estado='Terminado')
    prestamos_como_garante = Prestamo.objects.filter(garante=socio).exclude(estado='Terminado')

    if prestamos_como_titular.exists() or prestamos_como_garante.exists():
        messages.error(request, 'No se puede eliminar al socio. Tiene préstamos activos como titular o garante.')
        return redirect('socio_list')

    if request.method == 'POST':
        socio.activo = False  # o socio.delete() si deseas eliminar completamente
        socio.save()
        messages.success(request, 'Socio eliminado correctamente.')
        return redirect('socio_list')

    return render(request, 'socios/eliminar_socio.html', {'socio': socio})


@login_required
@role_required(allowed_roles=['Presidente', 'Secretaria'])
def ver_aportaciones_socio(request, socio_id):
    socio = get_object_or_404(Socio, pk=socio_id)
    movimientos = Movimiento.objects.filter(socio=socio).order_by('fecha_movimiento')

    total_aportes = movimientos.filter(salida=0).aggregate(total=Sum('entrada'))['total'] or 0
    total_retiros = movimientos.filter(entrada=0).aggregate(total=Sum('salida'))['total'] or 0
    saldo = total_aportes - total_retiros

    meses_con_aporte = movimientos.filter(entrada__gt=0).dates('fecha_movimiento', 'month')
    meses_con_aporte_set = set((m.year, m.month) for m in meses_con_aporte)

    meses_faltantes = []
    if movimientos.exists():
        inicio = movimientos.first().fecha_movimiento.replace(day=1)
        fin = date.today().replace(day=1)
        actual = inicio
        while actual <= fin:
            if (actual.year, actual.month) not in meses_con_aporte_set:
                meses_faltantes.append(actual.strftime('%B %Y'))
            actual += relativedelta(months=1)
            
    context = {
        'socio': socio,
        'movimientos': movimientos,
        'total_aportes': total_aportes,
        'total_retiros': total_retiros,
        'saldo': saldo,
        'meses_faltantes': meses_faltantes,
        'is_read_only': get_user_role(request.user) == 'Presidente', # Presidente no puede modificar aportes
    }
    return render(request, 'aportes/ver_aportaciones_socio.html', context)

@login_required
@role_required(allowed_roles=['Secretaria'])
def agregar_aporte(request, socio_id):
    socio = get_object_or_404(Socio, id=socio_id)

    if request.method == 'POST':
        detalle = request.POST.get('detalle_movimiento')
        tipo = request.POST.get('tipo')  # puede ser 'entrada' o 'salida'
        monto = request.POST.get('monto')
        fecha = request.POST.get('fecha_movimiento')

        if not monto:
            # Asegura que monto no sea None
            return HttpResponse("El campo monto es obligatorio.", status=400)

        monto_decimal = Decimal(monto)

        entrada = monto_decimal if tipo == 'entrada' else Decimal('0.00')
        salida = monto_decimal if tipo == 'salida' else Decimal('0.00')

        # Calcula el saldo actual del socio
        ultimo_mov = Movimiento.objects.filter(socio=socio).order_by('-fecha_movimiento').first()
        saldo_anterior = ultimo_mov.saldo if ultimo_mov else Decimal('0.00')
        nuevo_saldo = saldo_anterior + entrada - salida

        Movimiento.objects.create(
            socio=socio,
            detalle_movimiento=detalle,
            entrada=entrada,
            salida=salida,
            fecha_movimiento=fecha,
            saldo=nuevo_saldo,
        )

        return redirect('ver_aportaciones_socio', socio.id)


@login_required
@role_required(allowed_roles=['Secretaria'])
def editar_aporte(request, aporte_id):
    aporte = get_object_or_404(Movimiento, pk=aporte_id)
    if request.method == 'POST':
        detalle = request.POST.get('detalle_movimiento')
        tipo = request.POST.get('tipo')
        monto = request.POST.get('monto')
        fecha = request.POST.get('fecha_movimiento')

        try:
            monto_decimal = Decimal(monto)
        except (TypeError, ValueError):
            messages.error(request, "Monto inválido.")
            return redirect('ver_aportaciones_socio', socio_id=aporte.socio.id)

        aporte.detalle_movimiento = detalle
        aporte.fecha_movimiento = fecha
        if tipo == 'entrada':
            aporte.entrada = monto_decimal
            aporte.salida = Decimal('0.00')
        else:
            aporte.salida = monto_decimal
            aporte.entrada = Decimal('0.00')
        aporte.save()
        return redirect('ver_aportaciones_socio', socio_id=aporte.socio.id)
    return redirect('ver_aportaciones_socio', socio_id=aporte.socio.id)

@login_required
@role_required(allowed_roles=['Secretaria'])
def eliminar_aporte(request, aporte_id):
    aporte = get_object_or_404(Movimiento, pk=aporte_id)
    socio_id = aporte.socio.id
    if request.method == 'POST':
        aporte.delete()
        # Se debería recalcular el saldo de los movimientos posteriores
        return redirect('ver_aportaciones_socio', socio_id=socio_id)
    return render(request, 'eliminar_aporte_confirm.html', {'aporte': aporte})

# --- Vistas de Préstamos ---
@login_required
@role_required(allowed_roles=['Presidente', 'Tesorero'])
def prestamo_list(request):
    prestamos = Prestamo.objects.all()
    role = get_user_role(request.user)
    is_read_only = role == 'Presidente'
    context = {
        'prestamos': prestamos,
        'role': role,
        'is_read_only': is_read_only,
    }
    return render(request, 'prestamo/prestamo_list.html', context)

@login_required
@role_required(allowed_roles=['Tesorero'])
def crear_o_editar_prestamo(request, pk=None):
    prestamo = get_object_or_404(Prestamo, pk=pk) if pk else None
    config = Configuracion.objects.first()
    if request.method == 'POST':
        form = PrestamoForm(request.POST, instance=prestamo)
        if form.is_valid():
            prestamo = form.save(commit=False)
            prestamo.interes = config.tasa_interes
            if prestamo.plazo > config.plazo_maximo:
                form.add_error('plazo', f'El plazo máximo permitido es {config.plazo_maximo} meses.')
                return render(request, 'prestamo/crear_editar_prestamo.html', {'form': form})
            prestamo.estado = 'Solicitado' if not pk else 'Pendiente'
            prestamo.cuota = None
            prestamo.save()
            return redirect('prestamo_list')
    else:
        form = PrestamoForm(instance=prestamo)
        if prestamo:
            form.initial['cuota'] = ''
    return render(request, 'prestamo/crear_editar_prestamo.html', {'form': form})

def exportar_aportaciones_pdf(request, socio_id):
    socio = get_object_or_404(Socio, pk=socio_id)
    movimientos = Movimiento.objects.filter(socio=socio).order_by('fecha_movimiento')

    total_aportes = movimientos.filter(salida=0).aggregate(total=Sum('entrada'))['total'] or 0
    total_retiros = movimientos.filter(entrada=0).aggregate(total=Sum('salida'))['total'] or 0
    saldo = total_aportes - total_retiros

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="aportaciones_{socio.cedula}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    elements.append(Paragraph(f"Aportaciones del socio: <b>{socio.nombre} {socio.apellido}</b>", styles['Title']))
    elements.append(Paragraph(f"Cédula: {socio.cedula}", styles['Normal']))
    elements.append(Paragraph(f"Fecha de generación: {fecha_actual}", styles['Normal']))
    elements.append(Spacer(1, 12))

    data = [['Fecha', 'Descripción', 'Entrada', 'Salida']]
    for m in movimientos:
        data.append([
            m.fecha_movimiento.strftime('%d/%m/%Y'),
            m.detalle_movimiento,
            f"${m.entrada:,.2f}" if m.entrada else "",
            f"${m.salida:,.2f}" if m.salida else "",
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"<b>Total Aportes:</b> ${total_aportes:,.2f}", styles['Normal']))
    elements.append(Paragraph(f"<b>Total Retiros:</b> ${total_retiros:,.2f}", styles['Normal']))
    elements.append(Paragraph(f"<b>Saldo:</b> ${saldo:,.2f}", styles['Normal']))
    elements.append(Spacer(1, 36))
    elements.append(Paragraph("Firma Tesorero: _________________________", styles['Normal']))

    doc.build(elements)
    return response


@login_required
@role_required(allowed_roles=['Tesorero'])
@require_POST
def aprobar_prestamo(request, pk):
    prestamo = get_object_or_404(Prestamo, pk=pk)
    prestamo.estado = 'Aprobado'
    prestamo.fecha_aprobacion = date.today()
    prestamo.save()
    generar_amortizacion(prestamo)
    return redirect('prestamo_list')

@login_required
@role_required(allowed_roles=['Tesorero'])
@require_POST
def rechazar_prestamo(request, pk):
    prestamo = get_object_or_404(Prestamo, pk=pk)
    prestamo.estado = 'Rechazado'
    prestamo.save()
    return redirect('prestamo_list')


@login_required
def socio_list(request):
    filtro = request.GET.get('filtro', 'todos')
    socios = Socio.objects.filter(activo=True)  # 👈 Solo activos
    cargos = Cargo.objects.all()
    hoy = date.today().replace(day=1)

    def meses_faltantes(socio):
        movimientos = Movimiento.objects.filter(socio=socio, entrada__gt=0).order_by('fecha_movimiento')
        if not movimientos.exists():
            return True
        fechas_aportes = movimientos.dates('fecha_movimiento', 'month')
        meses_aporte_set = set((d.year, d.month) for d in fechas_aportes)
        inicio = movimientos.first().fecha_movimiento.replace(day=1)
        actual = inicio
        while actual <= hoy:
            if (actual.year, actual.month) not in meses_aporte_set:
                return True
            actual += relativedelta(months=1)
        return False

    socios_filtrados = []
    for socio in socios:
        faltante = meses_faltantes(socio)
        if filtro == 'al_dia' and not faltante:
            socios_filtrados.append(socio)
        elif filtro == 'deudores' and faltante:
            socios_filtrados.append(socio)
        elif filtro == 'todos':
            socios_filtrados.append(socio)
#paginacion
    paginator = Paginator(socios_filtrados, 10)
    page = request.GET.get('page')
    socios_paginados = paginator.get_page(page)

    return render(request, 'socios/socio_list.html', {
        'socios': socios_paginados,
        'filtro': filtro,
        'cargos': cargos,
    })


# Editar socio
def editar_socio(request, pk):
    socio = get_object_or_404(Socio, pk=pk)
    if request.method == 'POST':
        form = SocioForm(request.POST, request.FILES, instance=socio)
        if form.is_valid():
            form.save()
            return redirect('socio_list')
    else:
        form = SocioForm(instance=socio)

    return render(request, 'socios/crear_socio.html', {'form': form})



# Calcular totales por socio
def ver_aportaciones_socio(request, socio_id):
    socio = get_object_or_404(Socio, pk=socio_id)
    movimientos = Movimiento.objects.filter(socio=socio).order_by('fecha_movimiento')

    total_aportes = movimientos.filter(salida=0).aggregate(total=Sum('entrada'))['total'] or 0
    total_retiros = movimientos.filter(entrada=0).aggregate(total=Sum('salida'))['total'] or 0
    saldo = total_aportes - total_retiros

    # Detectar meses faltantes
    meses_con_aporte = movimientos.filter(entrada__gt=0).dates('fecha_movimiento', 'month')
    meses_con_aporte_set = set((m.year, m.month) for m in meses_con_aporte)

    meses_faltantes = []
    if movimientos.exists():
        inicio = movimientos.first().fecha_movimiento.replace(day=1)
        fin = date.today().replace(day=1)
        actual = inicio

        while actual <= fin:
            if (actual.year, actual.month) not in meses_con_aporte_set:
                meses_faltantes.append(actual.strftime('%B %Y'))
            actual += relativedelta(months=1)

    return render(request, 'aportes/ver_aportaciones_socio.html', {
        'socio': socio,
        'movimientos': movimientos,
        'total_aportes': total_aportes,
        'total_retiros': total_retiros,
        'saldo': saldo,
        'meses_faltantes': meses_faltantes,
    })


# Detalle del socio
def detalle_socio(request, pk):
    socio = get_object_or_404(Socio, pk=pk)
    # Calcular edad
    hoy = date.today()
    edad = hoy.year - socio.fecha_nacimiento.year - (
        (hoy.month, hoy.day) < (socio.fecha_nacimiento.month, socio.fecha_nacimiento.day)
    )

    return render(request, 'socios/detalle.html', {'socio': socio, 'edad': edad})


# Eliminar aporte
def eliminar_aporte(request, aporte_id):
    aporte = get_object_or_404(Movimiento, pk=aporte_id)
    socio_id = aporte.socio.id

    if request.method == 'POST':
        aporte.delete()
        return redirect('ver_aportaciones_socio', socio_id=socio_id)

    return render(request, 'eliminar_aporte_confirm.html', {'aporte': aporte})


@require_POST
def agregar_cargo(request):
    nombre = request.POST.get('nombre_cargo')
    estado = request.POST.get('estado') == 'true'
    Cargo.objects.create(nombre_cargo=nombre, estado=estado)
    return redirect('socio_list')

@require_POST
def editar_cargo(request, id):
    cargo = get_object_or_404(Cargo, id=id)
    cargo.nombre_cargo = request.POST.get('nombre_cargo')
    cargo.estado = request.POST.get('estado') == 'true'
    cargo.save()
    return redirect('socio_list')

@require_POST
def eliminar_cargo(request, id):
    cargo = get_object_or_404(Cargo, id=id)
    cargo.delete()
    return redirect('socio_list')

def prestamo_list(request):
    prestamos = Prestamo.objects.all()
    return render(request, 'prestamo/prestamo_list.html', {'prestamos': prestamos})

# Crear préstamo
def crear_o_editar_prestamo(request, pk=None):
    prestamo = get_object_or_404(Prestamo, pk=pk) if pk else None
    config = Configuracion.objects.first()  # Obtiene la configuración actual
    
    if request.method == 'POST':
        form = PrestamoForm(request.POST, instance=prestamo)
        if form.is_valid():
            prestamo = form.save(commit=False)

            # Asignar interés desde configuración (ignorar lo que venga en el form)
            prestamo.interes = config.tasa_interes
            
            # Validar plazo máximo
            if prestamo.plazo > config.plazo_maximo:
                form.add_error('plazo', f'El plazo máximo permitido es {config.plazo_maximo} meses.')
                return render(request, 'prestamo/crear_editar_prestamo.html', {'form': form})

            if not pk:
                prestamo.estado = 'Solicitado'
            else:
                prestamo.estado = 'Pendiente'

            prestamo.cuota = None
            prestamo.save()
            return redirect('prestamo_list')
    else:
        form = PrestamoForm(instance=prestamo)

        if prestamo:
            form.initial['cuota'] = ''

    return render(request, 'prestamo/crear_editar_prestamo.html', {'form': form})


#aprobar prestamo
@require_POST
def aprobar_prestamo(request, pk):
    prestamo = get_object_or_404(Prestamo, pk=pk)
    prestamo.estado = 'Aprobado'
    # Obtener la configuración activa
    config = Configuracion.objects.first()
    # Registrar gasto administrativo
    saldo_gasto = GastosAdministrativos.objects.order_by('-fecha').first()
    saldo_anterior_gasto = saldo_gasto.saldo if saldo_gasto else Decimal('0.00')
    nuevo_saldo_gasto = saldo_anterior_gasto + (prestamo.cantidad_aprobada * config.tasa_prestamo) / Decimal('100.00')
   # Crear gasto administrativo
    gasto = GastosAdministrativos(
        fecha=prestamo.fecha_aprobacion,
        descripcion=f'Tasa del {config.tasa_prestamo}% por préstamo aprobado de ' + str(prestamo.socio),
        entrada=(prestamo.cantidad_aprobada * config.tasa_prestamo) / Decimal('100.00'),
        salida=Decimal('0.00'),
        saldo=nuevo_saldo_gasto

    )
    gasto.save()

    #prestamo.fecha_aprobacion = date.today()
    prestamo.save()
    generar_amortizacion(prestamo)
    return redirect('prestamo_list')


#rechazar prestamo
@require_POST
def rechazar_prestamo(request, pk):
    prestamo = get_object_or_404(Prestamo, pk=pk)
    prestamo.estado = 'Rechazado'
    prestamo.save()
    return redirect('prestamo_list')


#pagos prestamos
@login_required
def pagos_prestamo(request, prestamo_id):
    prestamo = get_object_or_404(Prestamo, id=prestamo_id)
    pagos = PagoPrestamo.objects.filter(prestamo=prestamo).order_by('cuota_pago')

    return render(request, 'prestamo/pagos/pagos_prestamo.html', {
        'prestamo': prestamo,
        'pagos': pagos,
    })

#registros de pagos de  prestamos
@require_POST
@login_required
def registrar_pago(request, pago_id):
    pago = get_object_or_404(PagoPrestamo, id=pago_id)
    prestamo = pago.prestamo
# Captura datos del formulario modal
    fecha_pago = request.POST.get('fecha_pago')
    detalle_pago = request.POST.get('detalle_pago')
    comprobante = request.FILES.get('comprobante_pago')
    # Marcar el pago como realizado
    pago.estado = True
    pago.fecha_pago = fecha_pago or timezone.now().date()
    pago.detalle_pago = detalle_pago
    if comprobante:
        pago.comprobante_pago = comprobante
    
    pago.save()

    # Verificar si ya se completaron todos los pagos
    if not PagoPrestamo.objects.filter(prestamo=prestamo, estado=False).exists():
        # Cambiar el estado del préstamo a "Terminado"
        prestamo.estado = 'Terminado'
        prestamo.save()

        # Calcular el total de intereses generados
        total_interes = prestamo.pagos.aggregate(total=Sum('interes_pago'))['total'] or Decimal('0.00')

        # Obtener todos los socios activos
        socios = Socio.objects.filter(activo=True)
        num_socios = socios.count()

        if num_socios > 0 and total_interes > 0:
            interes_por_socio = total_interes / num_socios
            interes_por_socio = interes_por_socio.quantize(Decimal('0.01'))  # Redondear a 2 decimales

            for socio in socios:
                ultimo_mov = Movimiento.objects.filter(socio=socio).order_by('-fecha_movimiento').first()
                saldo_anterior = ultimo_mov.saldo if ultimo_mov else Decimal('0.00')
                nuevo_saldo = saldo_anterior + interes_por_socio

                Movimiento.objects.create(
                    socio=socio,
                    detalle_movimiento="Interes generado por prestamo",
                    entrada=interes_por_socio,
                    salida=Decimal('0.00'),
                    saldo=nuevo_saldo,
                    fecha_movimiento=timezone.now().date()
                )

    return redirect('pagos_prestamo', prestamo_id=prestamo.id)


#tabla amortizacion
def generar_amortizacion(prestamo):
    if prestamo.estado == 'Aprobado' and not prestamo.pagos.exists():
        capital_prestado = prestamo.cantidad_aprobada
        plazo = prestamo.plazo
        interes_anual = prestamo.interes
        saldo = capital_prestado

        # Tasa mensual en decimales
        tasa_mensual = (interes_anual / 100) / 12

        # Calcular cuota si no está definida
        if not prestamo.cuota:
            if tasa_mensual == 0:
                cuota = capital_prestado / plazo
            else:
                cuota = capital_prestado * (
                    tasa_mensual * (1 + tasa_mensual) ** plazo
                ) / ((1 + tasa_mensual) ** plazo - 1)
            prestamo.cuota = round(cuota, 2)
            prestamo.save()
        else:
            cuota = prestamo.cuota

        fecha = prestamo.fecha_aprobacion or timezone.now().date()

        for i in range(plazo):
            interes = saldo * tasa_mensual
            capital = cuota - interes

            # Redondeo para evitar errores por flotantes
            interes = round(interes, 2)
            capital = round(capital, 2)
            cuota_real = round(capital + interes, 2)

            # Última cuota ajustada
            if i == plazo - 1:
                capital = saldo
                interes = max(0, cuota - capital)
                cuota_real = round(capital + interes, 2)

            # Guardar cuota con saldo actual antes de restar
            PagoPrestamo.objects.create(
                prestamo=prestamo,
                cuota_pago=i + 1,
                saldo_pago=round(saldo, 2),
                capital_pago=capital,
                interes_pago=interes,
                plazo_pago=plazo - i,
                valor_cuota_pago=cuota_real,
                estado=False,
                fecha_a_pagar=fecha + relativedelta(months=i)
            )

            # Luego de registrar el pago, restamos el capital
            saldo -= capital
            saldo = round(saldo, 2)


def exportar_amortizacion_pdf(request, pk):
    prestamo = Prestamo.objects.get(pk=pk)
    pagos = prestamo.pagos.all()
    socio = prestamo.socio
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="amortizacion_prestamo_{pk}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    # Título principal

    elements.append(Paragraph(f'Tabla de Amortización - Préstamo #{pk}', styles['Title']))
    elements.append(Spacer(1, 12))
# Información del solicitante
    info_solicitante = f"""
        <strong>Solicitante:</strong> {socio.nombre} {socio.apellido}<br/>
        <strong>Fecha del Aprobación:</strong> {prestamo.fecha_aprobacion.strftime('%d/%m/%Y')}<br/>
        <strong>Plazo:</strong> {prestamo.plazo} meses<br/>
        <strong>Monto Aprobado:</strong> ${prestamo.cantidad_aprobada:,.2f}<br/>
        <strong>Interés:</strong> {prestamo.interes:,.2f}%
    """
    elements.append(Paragraph(info_solicitante, styles['Normal']))
    elements.append(Spacer(1, 12))
     # Cabecera de tabla
    datos = [['#', 'Saldo', 'Capital', 'Interés', 'Cuota', 'Fecha a Pagar', 'Estado']]

    for pago in pagos:
        datos.append([
            pago.cuota_pago,
            f"${pago.saldo_pago}",
            f"${pago.capital_pago}",
            f"${pago.interes_pago}",
            f"${pago.valor_cuota_pago}",
            pago.fecha_a_pagar.strftime('%d/%m/%Y'),
            "Pagado" if pago.estado else "Pendiente"
        ])

    tabla = Table(datos, repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#d3d3d3')),
        ('TEXTCOLOR',(0,0),(-1,0),colors.black),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
    ]))

    elements.append(tabla)
     # Firma
    elements.append(Spacer(1, 24))
    elements.append(Paragraph("__________________________", styles['Normal']))
    elements.append(Paragraph("Firma del Tesorero", styles['Normal']))
    doc.build(elements)
    return response

def exportar_amortizacion_excel(request, pk):
    prestamo = Prestamo.objects.get(pk=pk)
    pagos = prestamo.pagos.all()
    socio = prestamo.socio
    wb = Workbook()
    ws = wb.active
    ws.title = "Amortización"

    # 🔹 Información del solicitante
    ws.append(["Solicitante:", f"{socio.nombre} {socio.apellido}"])
    ws.append(["Fecha del Aprobación:", prestamo.fecha_aprobacion.strftime('%d/%m/%Y') if prestamo.fecha_aprobacion else ""])
    ws.append(["Plazo (meses):", prestamo.plazo])
    ws.append(["Monto Aprobado:", float(prestamo.cantidad_aprobada)])
    ws.append(["Interés (%):", float(prestamo.interes)])
    ws.append([])  # Fila vacía
     # 🔹 Cabecera de tabla
    headers = ['#', 'Saldo', 'Capital', 'Interés', 'Cuota', 'Fecha a Pagar', 'Estado']
    ws.append(headers)

    for pago in pagos:
        ws.append([
            pago.cuota_pago,
            pago.saldo_pago,
            pago.capital_pago,
            pago.interes_pago,
            pago.valor_cuota_pago,
            pago.fecha_a_pagar.strftime('%d/%m/%Y'),
            'Pagado' if pago.estado else 'Pendiente'
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="amortizacion_prestamo_{pk}.xlsx"'
    wb.save(response)
    return response

def exportar_socios_pdf(request):
    socios = Socio.objects.filter(activo=True).order_by('apellido')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="listado_socios.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # Título
    fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    elements.append(Paragraph("Listado de Socios Activos", styles['Title']))
    elements.append(Paragraph(f"📅 Generado el: {fecha_actual}", styles['Normal']))
    elements.append(Spacer(1, 12))

    # Cabecera de tabla
    datos = [['#', 'Cédula', 'Nombres', 'Apellidos', 'Teléfono', 'Email', 'Fecha de Ingreso']]

    # Filas de socios
    for i, socio in enumerate(socios, start=1):
        datos.append([
            i,
            socio.cedula,
            socio.nombre,
            socio.apellido,
            socio.telefono or '',
            socio.email or '',
            socio.fecha_ingreso.strftime('%d/%m/%Y')
        ])

    tabla = Table(datos, repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))

    elements.append(tabla)
     # Firma
    elements.append(Spacer(1, 24))
    elements.append(Paragraph("__________________________", styles['Normal']))
    elements.append(Paragraph("Firma del Secretario(a)", styles['Normal']))
    doc.build(elements)
    return response

def exportar_socios_excel(request):
    socios = Socio.objects.filter(activo=True).order_by('apellido')

    wb = Workbook()
    ws = wb.active
    ws.title = "Socios Activos"

    # Cabecera
    headers = ['#', 'Cédula', 'Nombres', 'Apellidos', 'Teléfono', 'Email', 'Fecha de Ingreso']
    ws.append(headers)

    # Estilo para la cabecera
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    # Contenido
    for i, socio in enumerate(socios, start=1):
        ws.append([
            i,
            socio.cedula,
            socio.nombre,
            socio.apellido,
            socio.telefono or '',
            socio.email or '',
            socio.fecha_ingreso.strftime('%d/%m/%Y')
        ])

    # Generar archivo
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="listado_socios.xlsx"'
    wb.save(response)
    return response

 # exportar listado de prestamos
def exportar_prestamos_pdf(request):
    prestamos = Prestamo.objects.all().order_by('-fecha_aprobacion')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="prestamos_lista.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    elements.append(Paragraph("📄 Lista de Préstamos", styles['Title']))
    elements.append(Paragraph(f"📅 Generado el: {fecha_actual}", styles['Normal']))
    elements.append(Spacer(1, 12))

    datos = [[
        'Fecha Aprob', 'Solicitante', 'Garante', 
        'Solicitado', 'Aprobado', 'Plazo', 'Estado'
    ]]

    for p in prestamos:
        datos.append([
            p.fecha_aprobacion.strftime('%d/%m/%Y') if p.fecha_aprobacion else '',
            f"{p.socio.nombre} {p.socio.apellido}",
            f"{p.garante.nombre} {p.garante.apellido}" if p.garante else '---',
            f"${p.cantidad_solicitada:,.2f}",
            f"${p.cantidad_aprobada:,.2f}" if p.cantidad_aprobada else '$0.00',
            f"{p.plazo} meses",
            p.estado
        ])

    tabla = Table(datos, repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
    ]))

    elements.append(tabla)
     # Firma
    elements.append(Spacer(1, 24))
    elements.append(Paragraph("__________________________", styles['Normal']))
    elements.append(Paragraph("Firma del Tesorero", styles['Normal']))
    doc.build(elements)
    return response

def exportar_prestamos_excel(request):
    prestamos = Prestamo.objects.all().order_by('-fecha_aprobacion')

    wb = Workbook()
    ws = wb.active
    ws.title = "Préstamos"

    headers = [
        'Fecha Aprobación', 'Solicitante', 'Garante', 
        'Cantidad Solicitada', 'Cantidad Aprobada', 'Plazo', 'Estado'
    ]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    for p in prestamos:
        ws.append([
            p.fecha_aprobacion.strftime('%d/%m/%Y') if p.fecha_aprobacion else '',
            f"{p.socio.nombre} {p.socio.apellido}",
            f"{p.garante.nombre} {p.garante.apellido}" if p.garante else '',
            float(p.cantidad_solicitada),
            float(p.cantidad_aprobada or 0),
            p.plazo,
            p.estado
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="prestamos_lista.xlsx"'
    wb.save(response)
    return response

def exportar_gastosadministrativos_pdf(request):
    gastos = GastosAdministrativos.objects.order_by('fecha')
    total_entrada = sum(g.entrada for g in gastos)
    total_salida = sum(g.salida for g in gastos)
    saldo_actual = total_entrada - total_salida

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="gastos_administrativos.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # Título y fecha
    fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    elements.append(Paragraph("📄 Reporte de Gastos Administrativos", styles['Title']))
    elements.append(Paragraph(f"📅 Generado el: {fecha_actual}", styles['Normal']))
    elements.append(Spacer(1, 12))

    # Tabla de datos
    data = [['Fecha', 'Descripción', 'Entrada', 'Salida']]
    for g in gastos:
        data.append([
            g.fecha.strftime('%d/%m/%Y'),
            g.descripcion,
            f"${g.entrada:,.2f}",
            f"${g.salida:,.2f}",
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 12))

    # Totales
    elements.append(Paragraph(f"<strong>Total Entradas:</strong> ${total_entrada:,.2f}", styles['Normal']))
    elements.append(Paragraph(f"<strong>Total Salidas:</strong> ${total_salida:,.2f}", styles['Normal']))
    elements.append(Paragraph(f"<strong>Saldo Actual:</strong> ${saldo_actual:,.2f}", styles['Normal']))
    elements.append(Spacer(1, 36))

    # Firma
    elements.append(Spacer(1, 24))
    elements.append(Paragraph("__________________________", styles['Normal']))
    elements.append(Paragraph("Firma del Tesorero", styles['Normal']))

    doc.build(elements)
    return response

def exportar_gastosadministrativos_excel(request):
    gastos = GastosAdministrativos.objects.order_by('fecha')
    total_entrada = sum(g.entrada for g in gastos)
    total_salida = sum(g.salida for g in gastos)
    saldo_actual = total_entrada - total_salida

    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos Administrativos"

    headers = ['Fecha', 'Descripción', 'Entrada', 'Salida', 'Saldo']
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    for g in gastos:
        ws.append([
            g.fecha.strftime('%d/%m/%Y'),
            g.descripcion,
            float(g.entrada),
            float(g.salida),
            float(g.saldo),
        ])

    # Totales al final
    ws.append([])
    ws.append(['', 'Total Entradas', float(total_entrada)])
    ws.append(['', 'Total Salidas', float(total_salida)])
    ws.append(['', 'Saldo Actual', float(saldo_actual)])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="gastos_administrativos.xlsx"'
    wb.save(response)
    return response

#dashboard
@login_required
def dashboard(request):
    socios_activos = Socio.objects.filter(activo=True).count()

    total_aportes = Movimiento.objects.filter(entrada__gt=0).aggregate(total=Sum('entrada'))['total'] or 0
    total_retiros = Movimiento.objects.filter(salida__gt=0).aggregate(total=Sum('salida'))['total'] or 0
    saldo_total = total_aportes - total_retiros

    # Totales por estado
    estados_deseados = ['Aprobado', 'Rechazado', 'Terminado']
    prestamos_estado_montos = Prestamo.objects.filter(estado__in=estados_deseados).values('estado').annotate(
        cantidad=Count('id'),
        monto=Sum('cantidad_solicitada')
    )

    # Para gráfico pastel
    total_aprobados = Prestamo.objects.filter(estado='Aprobado').aggregate(total=Sum('cantidad_solicitada'))['total'] or 0
    cartera_vencida = PagoPrestamo.objects.filter(estado=False).aggregate(
        total=Sum('saldo_pago'))['total'] or 0

    movimientos_por_mes = Movimiento.objects.annotate(
        mes=TruncMonth('fecha_movimiento')
    ).values('mes').annotate(
        total_entradas=Sum('entrada'),
        total_salidas=Sum('salida'),
    ).order_by('mes')

    prestamos_recientes = Prestamo.objects.order_by('-id')[:5]

    return render(request, 'dashboard.html', {
        'socios_activos': socios_activos,
        'total_aportes': total_aportes,
        'total_retiros': total_retiros,
        'saldo_total': saldo_total,
        'prestamos_estado_montos': prestamos_estado_montos,
        'total_aprobados': total_aprobados,
        'cartera_vencida': cartera_vencida,
        'movimientos_por_mes': movimientos_por_mes,
        'prestamos_recientes': prestamos_recientes,
    })           


def configuracion(request):
    config = Configuracion.objects.first()
    if request.method == 'POST':
        form = ConfiguracionForm(request.POST, request.FILES, instance=config)
        if form.is_valid():
            form.save()
            return redirect('dashboard')
    else:
        form = ConfiguracionForm(instance=config)
    return render(request, 'configuracion/configuracion.html', {'form': form, 'configuracion': config})


#Gastos administrativos
def gastos_administrativos(request, action=None, pk=None):
    if action == 'agregar':
        form = GastoAdministrativoForm(request.POST or None)
        if request.method == 'POST' and form.is_valid():
            nuevo = form.save(commit=False)
            ultimo = GastosAdministrativos.objects.order_by('-fecha').first()
            saldo_anterior = ultimo.saldo if ultimo else Decimal('0.00')
            nuevo.saldo = saldo_anterior + nuevo.entrada - nuevo.salida
            nuevo.save()
            return redirect('gastos_admin')
        return render(request, 'gastos/form_gasto_admin.html', {
            'form': form,
            'modo': 'Agregar'
        })

    elif action == 'editar' and pk:
        gasto = get_object_or_404(GastosAdministrativos, pk=pk)
        form = GastoAdministrativoForm(request.POST or None, instance=gasto)
        if request.method == 'POST' and form.is_valid():
            gasto = form.save(commit=False)
            # Aquí podrías recalcular saldo si es necesario, o mantenerlo igual
            gasto.save()
            return redirect('gastos_admin')
        return render(request, 'gastos/form_gasto_admin.html', {
            'form': form,
            'modo': 'Editar'
        })

    elif action == 'eliminar' and pk:
        gasto = get_object_or_404(GastosAdministrativos, pk=pk)
        if request.method == 'POST':
            gasto.delete()
            return redirect('gastos_admin')
        return render(request, 'gastos/eliminar_gasto_admin.html', {'gasto': gasto})

    # Lista de gastos por defecto
    gastos = GastosAdministrativos.objects.exclude(id__isnull=True)
    total_entrada = gastos.aggregate(total=Sum('entrada'))['total'] or 0
    total_salida = gastos.aggregate(total=Sum('salida'))['total'] or 0
    saldo_actual = total_entrada - total_salida

    return render(request, 'gastos/gastos_admin_list.html', {
        'gastos': gastos,
        'total_entrada': total_entrada,
        'total_salida': total_salida,
        'saldo_actual': saldo_actual
    })
def imprimir_socios(request):
    socios = Socio.objects.filter(activo=True).order_by('apellido')
    return render(request, 'socios/imprimir/socios_imprimir.html', {'socios': socios})